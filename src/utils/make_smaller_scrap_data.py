from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
raw_data_dir = '../../raw_data/'

# read autohome scrap
scrap_workbook = load_workbook(filename = raw_data_dir + "autohome_scrap.xlsx", read_only = True, data_only = True)

# ES6
ES6_sheet = scrap_workbook["蔚来ES6"]
output_workbook = Workbook()
output_sheet = output_workbook.create_sheet("蔚来ES6", 0)

for review in ES6_sheet.iter_rows(min_row = 1, max_row = 12, values_only = True):
    output_sheet.append(review)

output_workbook.save(filename = raw_data_dir + "smaller_es6_autohome_scrap.xlsx")
