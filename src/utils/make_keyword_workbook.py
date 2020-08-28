from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
raw_data_dir = '../../raw_data/'

# read suggested keyword
keyword_workbook = load_workbook(filename = "../../keyword.xlsx", read_only = True, data_only = True)
suggested_keyword_sheet = keyword_workbook["关键词组"]
output_workbook = Workbook()
output_sheet = output_workbook.create_sheet("关键词组", 0)
for rows in suggested_keyword_sheet.iter_rows(min_row = 3, min_col = 2, values_only = True):
    output_sheet.append(rows)

output_workbook.save(filename = raw_data_dir + "smaller_keyword.xlsx")

