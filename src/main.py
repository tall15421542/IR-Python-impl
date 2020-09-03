import argparse
import os

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import jieba
import pandas as pd

# custom import
from inverted_file import InvertedFile, InvertedIndex, TermDocInfo, DocWordCountInfo  
from term_container import TermContainer
from query import Query, QueryExpandImpl 
from review import ReviewContainer
from tfidf import TFIDF

# read autohome scrap
def read_scrap(filename):
    return load_workbook(filename = filename, read_only = True, data_only = True)

def get_review_list(sheet):
    review_list = []
    for col in ES6_sheet.iter_rows(min_row = 2, min_col = 8, max_col = 8, values_only = True):
        review = col[0]
        review_list.append(review)
    return review_list

def build_doc_word_count_info_list(review_list):
    doc_word_count_info_list = []
    for doc_id, review in enumerate(review_list):
        doc_word_count_info_list.append(DocWordCountInfo(doc_id, review))
    return doc_word_count_info_list

def get_query_list(workbook_path):
    query_list = []
    query_expand_workbook = load_workbook(filename = workbook_path, read_only = True, data_only = True)
    query_expand_sheet = query_expand_workbook["关键词组"]
    for rows in query_expand_sheet.iter_rows(values_only = True):
        query = rows[0]
        if query:
            query = query.replace(" ", "")
            query_list.append(Query(query))
    return query_list

def apply_query_expand_to_query_list(query_list, query_expand_impl):
    for query in query_list:
        query.query_expand(query_expand_impl)
        query.preprocess()

def apply_query_search_to_query_list(query_list, inverted_file, tfidf_engine, review_container):
    for query in query_list:
        print("Query: ", query.get_query())
        print("----------------")
        query.search_and_update_topk_review(inverted_file, tfidf_engine, review_container)

def set_topk_for_query_list(query_list, topk):
    for query in query_list:
        query.set_topk(topk)

def update_workbook_for_query_list(query_list, review_container, workbook):
    for query in query_list:
        query.update_workbook_sheet(review_container, workbook)

if __name__ == "__main__":
    # argument list
    parser = argparse.ArgumentParser(description = "Scrap_search_engine")
    parser.add_argument('-i', action = 'store', dest = 'scrap_file_name', required = True)
    parser.add_argument('-f', action = 'store', dest = 'query_expand_workbook_path')
    parser.add_argument('-k', action = 'store', dest = 'topk', type = int, required = True)
    parser.add_argument('-o', action = 'store', dest = 'output_path', default = "../output/search_result.xlsx")
    args = parser.parse_args()
    
    # read scrap_workbook
    scrap_workbook = read_scrap(args.scrap_file_name)
    #
    ## ES6
    ES6_sheet = scrap_workbook["蔚来ES6"]
    review_container = ReviewContainer(ES6_sheet)
    review_list = review_container.get_review_list()
    doc_word_count_info_list = build_doc_word_count_info_list(review_list)
    
    ## build model data structure
    term_container = TermContainer(doc_word_count_info_list)
    inverted_file = InvertedFile(term_container, doc_word_count_info_list)
    
    # build query
    query_list = get_query_list(args.query_expand_workbook_path)
    query_expand_impl = QueryExpandImpl(args.query_expand_workbook_path)
    set_topk_for_query_list(query_list, args.topk)
    apply_query_expand_to_query_list(query_list,  query_expand_impl)

    # search 
    tfidf_engine = TFIDF(review_container)    
    apply_query_search_to_query_list(query_list, inverted_file, tfidf_engine, review_container)  
   
    # output_workbook
    workbook = Workbook()
    update_workbook_for_query_list(query_list, review_container, workbook)
    workbook.remove(workbook['Sheet'])
    workbook.save(args.output_path)
