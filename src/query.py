from openpyxl import load_workbook
from openpyxl import Workbook
import heapq 
import jieba

from review import ReviewAccumulatorDict

class Query:
    def __init__(self, query):
        self.query = query
        self.query_expansion_list = []
        self.term_freq_dict = {}
        self.top_k = 2
        self.topk_review_queue = []
        heapq.heapify(self.topk_review_queue)

    def query_expand(self, query_expand_impl):
        self.query_expansion_list = query_expand_impl.expand_for_search(self.query)

    def get_query(self):
        return self.query

    def print_query_expansion(self):
        print(self.query_expansion_list)

    def preprocess(self):
        for term in self.query_expansion_list:
            tf = self.term_freq_dict.get(term, 0)
            tf = tf + 1
            self.term_freq_dict[term] = tf

    def print_term_freq_dict(self):
        for term, tf in self.term_freq_dict.items():
            print(term, tf)

    def search_and_update_topk_review(self, inverted_file, tfidf_engine, review_container):
        review_accumulator_dict = ReviewAccumulatorDict(inverted_file, tfidf_engine, review_container)
        for term in self.query_expansion_list:
            qtf = self.term_freq_dict[term]
            review_accumulator_dict.update(qtf, term)
        self.update_topk_review_queue(review_accumulator_dict)

    def update_topk_review_queue(self, review_accumulator_dict):
        print(self.top_k)
        for review_accumulator in review_accumulator_dict.values():
            if len(self.topk_review_queue) < self.top_k:
                heapq.heappush(self.topk_review_queue, review_accumulator)
            elif self.topk_review_queue[0] < review_accumulator:
                heapq.heappop(self.topk_review_queue)
                heapq.heappush(self.topk_review_queue, review_accumulator)
        self.topk_review_queue = heapq.nlargest(self.top_k, self.topk_review_queue)

    def update_workbook_sheet(self, review_container, workbook):
        sheet = workbook.create_sheet(self.query)
        sheet.append(["review", "score"])
        for review_accumulator in self.topk_review_queue:
            review = review_container.get_review(review_accumulator.doc_id)
            score = review_accumulator.get_score()
            sheet.append([review, score])
    
    def set_topk(self, topk):
        self.top_k = topk

class QueryExpandImpl:
    def __init__(self, query_expand_workbook_path):
        self.query_expand_backend_dict = {}
        self.build_backend(query_expand_workbook_path)

    def build_backend(self, query_expand_workbook_path):
        query_expand_workbook = load_workbook(filename = query_expand_workbook_path, read_only = True, data_only = True)
        query_expand_sheet = query_expand_workbook["关键词组"]
        for rows in query_expand_sheet.iter_rows(values_only = True):
            query = rows[0]
            if query:
                query = query.replace(" ", "")
                query_expansion_list = self.cell_tuple_to_list(rows)
                self.query_expand_backend_dict[query] = query_expansion_list
    
    def cell_tuple_to_list(self, cell_tuple):
        cell_list = []
        for idx, cell in enumerate(cell_tuple):
            if idx == 0:
                continue
            if cell:
                cell = cell.replace(" ", "")
                cell_list.append(cell.upper())
        return cell_list

    def expand_for_search(self, query):
        if query in self.query_expand_backend_dict:
            query_expand_list = self.query_expand_backend_dict[query]
            expand_for_search_list = []
            for term in query_expand_list:
                seg_list = jieba.cut_for_search(term)
                expand_for_search_list.extend(list(jieba.cut_for_search(term)))
            return expand_for_search_list
        return query

   

    
    

        
